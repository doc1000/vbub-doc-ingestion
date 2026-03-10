"""XLSX extractor using openpyxl.

Loads the workbook in read-only, data-only mode.  For each sheet:
  - Reads all rows as string lists (None cells become empty strings).
  - Applies the tabular text policy (numeric suppression).
  - Flattens rows to pipe-delimited lines.
  - Prepends a "## Sheet: <name>" label.

Sheets are joined with double newlines.
Title comes from the first non-empty cell of the first sheet's header row.
"""

import io

import openpyxl

from ingestion_service.app.domain.contracts import ExtractionResult
from ingestion_service.app.services.tabular_text_policy_service import filter_rows

_PARSER_NAME = "XlsxExtractor"
_PARSER_VERSION = "0.1.0"
_CELL_SEPARATOR = " | "


def _cell_to_str(value: object) -> str:
    """Convert an openpyxl cell value to a clean string."""
    if value is None:
        return ""
    return str(value).strip()


def _flatten_rows(rows: list[list[str]]) -> str:
    """Join each row's cells with ' | ' and join rows with newlines."""
    return "\n".join(_CELL_SEPARATOR.join(row) for row in rows)


class XlsxExtractor:
    """Extracts readable text from XLSX files using openpyxl.

    Implements the DocumentExtractor Protocol.
    """

    def extract(
        self,
        file_bytes: bytes,
        filename: str,
        canonical_mime: str,
    ) -> ExtractionResult:
        """Load, parse, filter, and flatten an XLSX workbook to text.

        Each sheet produces a labeled block of pipe-delimited rows.
        Numeric-heavy rows are suppressed by the tabular text policy.

        Args:
            file_bytes:     raw bytes of the XLSX file.
            filename:       original filename (informational).
            canonical_mime: not used for dispatch here.

        Returns:
            ExtractionResult with sheet-labeled pipe-delimited clean_text
            and an optional title from the first sheet's header.
        """
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_bytes), data_only=True, read_only=True
        )

        sheet_blocks: list[str] = []
        title: str | None = None

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows: list[list[str]] = [
                [_cell_to_str(cell.value) for cell in row]
                for row in sheet.iter_rows()
                if any(cell.value is not None for cell in row)
            ]

            filtered = filter_rows(rows, source_type="xlsx")
            if not filtered:
                continue

            # Extract title from the first sheet's first non-empty header cell.
            if title is None and filtered[0]:
                candidate = filtered[0][0].strip()
                if candidate:
                    title = candidate

            flat = _flatten_rows(filtered)
            sheet_blocks.append(f"## Sheet: {sheet_name}\n{flat}")

        workbook.close()

        clean_text = "\n\n".join(sheet_blocks)

        return ExtractionResult(
            parser_name=_PARSER_NAME,
            parser_version=_PARSER_VERSION,
            title=title,
            clean_text=clean_text,
            warnings=[],
        )
